from flask import Blueprint, render_template, session, redirect, url_for, request, flash
from app.db import get_db_connection
from psycopg import errors
import openpyxl

managers_bp = Blueprint("managers", __name__, url_prefix="/managers")

# A6: Managers Overview (read-only for any logged-in user)
@managers_bp.route("/overview")
def managers_overview():
    if "user_id" not in session:
        return redirect(url_for("auth.login"))
    
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Query for department overview with manager info and stats
        cur.execute("""
            SELECT 
                d.dname AS department_name,
                d.dnumber AS department_number,
                COALESCE(e.fname || ' ' || e.minit || '. ' || e.lname, 'N/A') AS manager_name,
                COUNT(DISTINCT emp.ssn) AS employee_count,
                COALESCE(SUM(w.hours), 0) AS total_hours
            FROM department d
            LEFT JOIN employee e ON d.mgr_ssn = e.ssn
            LEFT JOIN employee emp ON d.dnumber = emp.dno
            LEFT JOIN works_on w ON emp.ssn = w.essn
            GROUP BY d.dnumber, d.dname, e.fname, e.minit, e.lname
            ORDER BY d.dname
        """)
        departments = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return render_template("managers.html", departments=departments)
    
    except Exception:
        return render_template("managers.html", departments=[])

# Excel Import Bonus Feature (admin only)
@managers_bp.route("/import", methods=["GET"])
def import_departments_form():
    if "user_id" not in session:
        return redirect(url_for("auth.login"))
    if session.get("role") != "admin":
        flash("You do not have permission to import departments.", "error")
        return redirect(url_for("managers.managers_overview"))
    
    return render_template("import_departments.html")

@managers_bp.route("/import", methods=["POST"])
def import_departments():
    if "user_id" not in session:
        return redirect(url_for("auth.login"))
    if session.get("role") != "admin":
        flash("You do not have permission to import departments.", "error")
        return redirect(url_for("managers.managers_overview"))
    
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(request.url)
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(request.url)
    
    if not file.filename.endswith('.xlsx'):
        flash('Only .xlsx files are allowed', 'error')
        return redirect(request.url)
    
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        success_count = 0
        error_rows = []
        
        # Skip header row (row 1), start from row 2
        for row_num in range(2, sheet.max_row + 1):
            dnumber = sheet.cell(row=row_num, column=1).value
            dname = sheet.cell(row=row_num, column=2).value
            mgr_ssn = sheet.cell(row=row_num, column=3).value
            
            # Skip empty rows
            if dnumber is None and dname is None:
                continue
                
            try:
                # Validate required fields
                if dnumber is None or dname is None:
                    error_rows.append(f"Row {row_num}: Missing required fields")
                    continue
                
                # Validate data types
                try:
                    dnumber = int(dnumber)
                except (ValueError, TypeError):
                    error_rows.append(f"Row {row_num}: Department number must be a number")
                    continue
                
                # Insert into department table
                cur.execute("""
                    INSERT INTO department (dnumber, dname, mgr_ssn)
                    VALUES (%s, %s, %s)
                """, (dnumber, dname, mgr_ssn))
                
                success_count += 1
                
            except errors.UniqueViolation:
                error_rows.append(f"Row {row_num}: Department number {dnumber} already exists")
                conn.rollback()
                continue
            except errors.ForeignKeyViolation:
                error_rows.append(f"Row {row_num}: Manager SSN {mgr_ssn} not found in employees")
                conn.rollback()
                continue
            except Exception as e:
                error_rows.append(f"Row {row_num}: {str(e)}")
                conn.rollback()
                continue
        
        if success_count > 0:
            conn.commit()
            flash(f'Successfully imported {success_count} departments', 'success')
        
        if error_rows:
            error_message = " | ".join(error_rows[:5])
            if len(error_rows) > 5:
                error_message += f" ... and {len(error_rows) - 5} more errors"
            flash(f'Import errors: {error_message}', 'error')
        
        cur.close()
        conn.close()
        
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'error')
    
    return redirect(url_for('managers.import_departments_form'))
